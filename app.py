# -*- coding: utf-8 -*-
import os
import re
import uuid
import logging
import magic
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import datetime # برای context processor
from typing import List, Dict, Tuple, Optional, Set, Union, cast
import pandas as pd
from flask import Flask, render_template, request, send_file, Response
from werkzeug.utils import secure_filename

# --- Selenium Imports ---
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, WebDriverException, NoSuchElementException,
    ElementNotVisibleException, SessionNotCreatedException
)
# --- Webdriver Manager Import ---
from webdriver_manager.chrome import ChromeDriverManager # برای مدیریت خودکار درایور

# ------------------------------------

from email_validator import validate_email, EmailNotValidError
from dotenv import load_dotenv # برای خواندن .env در توسعه محلی

# --- Load Environment Variables ---
# این فایل .env معمولا در production استفاده نمی‌شود، متغیرها مستقیم در محیط تنظیم می‌شوند
load_dotenv()

# --- Configuration ---
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
# پوشه instance برای فایل‌های آپلودی، نتایج و لاگ‌ها - باید توسط سرور WSGI قابل نوشتن باشد
INSTANCE_FOLDER = os.path.join(BASE_DIR, 'instance')
UPLOAD_FOLDER = os.path.join(INSTANCE_FOLDER, 'uploads')
RESULTS_FOLDER = os.path.join(INSTANCE_FOLDER, 'results')
LOG_FILE = os.path.join(INSTANCE_FOLDER, 'app.log')
# مسیر فایل‌های stopwords - فرض می‌شود در کنار فایل اصلی و در ریپازیتوری هستند
STOPWORDS_DIR = BASE_DIR

ALLOWED_EXTENSIONS: Set[str] = {'.xlsx'}
SELENIUM_TIMEOUT = 30 # ثانیه
MAX_FILE_SIZE = 10 * 1024 * 1024 # 10 مگابایت
SNIPPET_CONTEXT_LENGTH = 50 # تعداد کاراکتر اطراف کلمه کلیدی در پیش‌نمایش

# --- Helper Function to Load Stop Words ---
def load_stopwords_from_file(filepath: str) -> Set[str]:
    """Loads stop words from a file (one word per line)."""
    try:
        # اطمینان از وجود فایل قبل از باز کردن
        if not os.path.isfile(filepath):
             logging.warning(f"Stopwords file not found: {filepath}. Returning empty set.")
             return set()
        with open(filepath, 'r', encoding='utf-8') as f:
            stopwords = {line.strip().lower() for line in f if line.strip()}
        logging.info(f"Successfully loaded {len(stopwords)} stopwords from {filepath}")
        return stopwords
    except Exception as e:
        logging.error(f"Error reading stopwords file {filepath}: {e}", exc_info=True)
        return set() # برگرداندن مجموعه خالی در صورت خطا

# --- Load Stop Words ---
# این فایل‌ها باید در ریپازیتوری شما وجود داشته باشند
STOP_WORDS_ENGLISH: Set[str] = load_stopwords_from_file(os.path.join(STOPWORDS_DIR, 'stopwords_en.txt'))
STOP_WORDS_PERSIAN: Set[str] = load_stopwords_from_file(os.path.join(STOPWORDS_DIR, 'stopwords_fa.txt'))
STOP_WORDS_ARABIC: Set[str] = load_stopwords_from_file(os.path.join(STOPWORDS_DIR, 'stopwords_ar.txt'))
ALL_STOP_WORDS: Set[str] = STOP_WORDS_ENGLISH.union(STOP_WORDS_PERSIAN).union(STOP_WORDS_ARABIC)

# --- Folder Setup ---
# اطمینان از وجود پوشه‌ها در زمان اجرای برنامه
# در محیط production، این پوشه‌ها باید توسط یوزری که برنامه را اجرا می‌کند قابل نوشتن باشند
for folder in [UPLOAD_FOLDER, RESULTS_FOLDER]:
    try:
        os.makedirs(folder, exist_ok=True)
    except OSError as e:
        logging.error(f"Could not create folder {folder}: {e}")
        # در صورت عدم توانایی ساخت پوشه، ممکن است لازم باشد برنامه متوقف شود یا خطا دهد
        # raise RuntimeError(f"Required folder {folder} could not be created.") from e

# --- Flask App ---
app = Flask(__name__, instance_path=INSTANCE_FOLDER)
app.config.update(
    UPLOAD_FOLDER=UPLOAD_FOLDER,
    RESULTS_FOLDER=RESULTS_FOLDER,
    MAX_CONTENT_LENGTH=MAX_FILE_SIZE,
    # !!! بسیار مهم: کلید مخفی در production هرگز نباید مقدار پیش‌فرض ضعیف باشد !!!
    # !!! این متغیر باید حتماً در محیط سرور (Environment Variable) تنظیم شود !!!
    SECRET_KEY=os.environ.get('FLASK_SECRET_KEY', 'change_this_in_production_env_variable!')
)

# بررسی اینکه آیا کلید مخفی مقدار پیش‌فرض دارد (فقط برای هشدار در لاگ‌ها)
if app.config['SECRET_KEY'] == 'change_this_in_production_env_variable!':
    logging.warning("SECURITY WARNING: Using default SECRET_KEY. Set the FLASK_SECRET_KEY environment variable in production!")

# --- Logging ---
# تنظیم لاگ‌گیری - در production ممکن است بخواهید از لاگ‌های چرخشی یا سیستم‌های متمرکز استفاده کنید
logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO, # در production ممکن است بخواهید سطح را به WARNING یا ERROR تغییر دهید
    format='%(asctime)s - %(levelname)s - %(process)d - %(thread)d - %(message)s'
)
logging.getLogger('selenium').setLevel(logging.WARNING)
logging.getLogger('urllib3').setLevel(logging.WARNING)
logging.getLogger('WDM').setLevel(logging.WARNING) # لاگ‌های webdriver-manager

# --- Jinja Context Processor ---
@app.context_processor
def inject_now():
    """Injects current time (UTC) into templates."""
    return {'now': datetime.datetime.utcnow()}

# --- Helper Functions --- (توابع کمکی بدون تغییر نسبت به نسخه قبل)

def allowed_file(filename: str) -> bool:
    """Checks if the file extension is allowed."""
    return '.' in filename and os.path.splitext(filename)[1].lower() in ALLOWED_EXTENSIONS

def is_valid_excel(filepath: str) -> bool:
    """Validates if the file is an actual Excel file using MIME type."""
    try:
        mime = magic.Magic(mime=True)
        mime_type = mime.from_file(filepath)
        return mime_type in [
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ]
    except Exception as e:
        logging.error(f"Magic library error checking file {filepath}: {e}")
        return os.path.splitext(filepath)[1].lower() in ALLOWED_EXTENSIONS

def validate_url(url: str) -> str:
    """Validates the URL and ensures it starts with http:// or https://"""
    url = url.strip()
    if not re.match(r'^(http|https)://', url):
         logging.info(f"URL '{url}' did not start with http/https. Prepending 'https://'.")
         return 'https://' + url
    return url

def validate_email_address(email: str) -> bool:
    """Validates the email format."""
    try:
        validate_email(email, check_deliverability=False) # عدم بررسی تحویل‌پذیری برای سرعت
        return True
    except EmailNotValidError as e:
        logging.warning(f"Invalid email format: {email} - {e}")
        return False

def fetch_page_text(url: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Fetches and cleans text content from a URL using Selenium with webdriver-manager.
    Returns (page_text, error_message)
    """
    driver = None
    logging.info(f"Attempting to fetch URL using Selenium: {url}")
    try:
        chrome_options = Options()
        chrome_options.add_argument('--headless=new') # حالت headless جدید کروم
        chrome_options.add_argument('--no-sandbox') # ضروری در محیط‌های لینوکسی/کانتینری
        chrome_options.add_argument('--disable-dev-shm-usage') # جلوگیری از مشکلات حافظه اشتراکی در داکر/لینوکس
        chrome_options.add_argument('--disable-gpu') # معمولا برای headless لازم است
        # استفاده از User-Agent عمومی برای جلوگیری از بلاک شدن
        chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36")
        chrome_options.add_argument('--blink-settings=imagesEnabled=false') # غیرفعال کردن تصاویر برای سرعت
        chrome_options.add_argument("--window-size=1920,1080")
        chrome_options.add_argument("--lang=en-US,en;q=0.9") # تنظیم زبان درخواست‌ها
        # chrome_options.add_argument("--disable-extensions")
        # chrome_options.add_argument("--proxy-server='direct://'") # عدم استفاده از پروکسی
        # chrome_options.add_argument("--proxy-bypass-list=*")
        # chrome_options.add_argument("--start-maximized")

        logging.info("Initializing Chrome driver using webdriver-manager...")
        # نصب یا به‌روزرسانی خودکار درایور کروم
        service = ChromeService(ChromeDriverManager().install())
        driver = Chrome(service=service, options=chrome_options)
        logging.info("Chrome driver initialized successfully.")

        driver.set_page_load_timeout(SELENIUM_TIMEOUT)
        driver.set_script_timeout(SELENIUM_TIMEOUT)

        logging.info(f"Navigating to {url}")
        driver.get(url)

        WebDriverWait(driver, SELENIUM_TIMEOUT).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )

        html = driver.page_source
        if not html:
            logging.warning(f"No HTML content retrieved from {url}")
            return None, "محتوایی از صفحه دریافت نشد."

        # --- Text Cleaning ---
        text = re.sub(r'<script[^>]*>.*?</script>', ' ', html, flags=re.IGNORECASE | re.DOTALL)
        text = re.sub(r'<style[^>]*>.*?</style>', ' ', text, flags=re.IGNORECASE | re.DOTALL)
        text = re.sub(r'<[^>]+>', ' ', text)
        text = re.sub(r'\s+', ' ', text).strip().lower()

        logging.info(f"Successfully fetched and cleaned text from {url}. Length: {len(text)}")
        return text, None

    # --- Error Handling --- (بدون تغییر نسبت به نسخه قبل)
    except TimeoutException as e:
        logging.error(f"Timeout error fetching {url}: {e}", exc_info=True)
        return None, f"خطای وقفه زمانی (Timeout) هنگام بارگذاری صفحه: {url}"
    except SessionNotCreatedException as e:
         logging.error(f"SessionNotCreatedException fetching {url}: {e}", exc_info=True)
         return None, "خطا در ایجاد نشست مرورگر: ممکن است نسخه ChromeDriver با Chrome سازگار نباشد یا مشکلی در نصب خودکار درایور وجود داشته باشد. لطفاً از به‌روز بودن Chrome روی سرور اطمینان حاصل کنید."
    except WebDriverException as e:
        logging.error(f"WebDriver error fetching {url}: {e}", exc_info=True)
        if "net::ERR_NAME_NOT_RESOLVED" in str(e) or "dns error" in str(e).lower():
            return None, f"آدرس وب‌سایت نامعتبر یا در دسترس نیست: {url}"
        if "unable to connect" in str(e).lower():
            return None, f"امکان برقراری ارتباط با وب‌سایت وجود ندارد: {url}"
        if "failed to start" in str(e).lower() or "terminated unexpectedly" in str(e).lower():
             return None, "خطا در اجرای ChromeDriver. ممکن است مشکلی در نصب خودکار یا مجوزهای اجرایی وجود داشته باشد."
        # خطاهای مربوط به دسترسی به حافظه یا مشکلات دیگر درایور
        if "DevToolsActivePort file doesn't exist" in str(e):
             logging.error("DevToolsActivePort error - potentially resource exhaustion or browser crash.")
             return None, "خطا در ارتباط با مرورگر (DevTools). ممکن است منابع سرور کم باشد یا مرورگر کرش کرده باشد."
        return None, f"خطای مرورگر (WebDriver) هنگام دسترسی به صفحه: {e}"
    except (NoSuchElementException, ElementNotVisibleException) as e:
        logging.error(f"Element error on {url}: {e}", exc_info=True)
        return None, f"خطا در یافتن المان‌های ضروری صفحه: {e}"
    except Exception as e:
        logging.error(f"Unexpected error fetching {url}: {e}", exc_info=True)
        return None, f"خطای پیش‌بینی نشده در دریافت اطلاعات صفحه: {e}"
    finally:
        if driver:
            try:
                driver.quit()
                logging.info(f"Selenium driver quit for {url}")
            except Exception as e:
                logging.error(f"Error quitting driver for {url}: {e}")


def get_phrases_from_file(filepath: str) -> Tuple[Optional[List[str]], Optional[str]]:
    """Reads phrases (one per row) from the first column of an Excel file."""
    try:
        df = pd.read_excel(filepath, header=None, engine='openpyxl')
        if df.empty or df.shape[1] == 0:
            return None, "فایل اکسل خالی است یا ستون اول وجود ندارد."
        phrases = df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
        valid_phrases = [p for p in phrases if p]
        if not valid_phrases:
             return None, "هیچ عبارت معتبری در ستون اول فایل اکسل یافت نشد."
        logging.info(f"Read {len(valid_phrases)} phrases from {filepath}")
        return valid_phrases, None
    except FileNotFoundError:
        logging.error(f"Excel file not found at path: {filepath}")
        return None, f"فایل اکسل یافت نشد: {filepath}"
    except pd.errors.EmptyDataError:
        logging.warning(f"Excel file is empty: {filepath}")
        return None, "فایل اکسل خالی است."
    except ImportError:
        logging.error("Pandas 'openpyxl' engine not installed. Run: pip install openpyxl")
        return None, "وابستگی 'openpyxl' برای خواندن فایل‌های .xlsx نصب نشده است."
    except Exception as e:
        logging.error(f"Error reading Excel file {filepath}: {e}", exc_info=True)
        if "No sheet named" in str(e):
             return None, f"خطا در خواندن فایل اکسل: شیت مورد نظر یافت نشد. ({e})"
        return None, f"خطا در خواندن یا پردازش فایل اکسل: {e}"


def get_context_snippet(text: str, phrase: str, context_len: int = SNIPPET_CONTEXT_LENGTH) -> Optional[str]:
    """Extracts a snippet of text around the first occurrence of the phrase (case-insensitive)."""
    try:
        matches = list(re.finditer(re.escape(phrase), text, re.IGNORECASE))
        if not matches:
            return None
        match = matches[0]
        start, end = match.span()
        snippet_start = max(0, start - context_len)
        snippet_end = min(len(text), end + context_len)
        prefix = "..." if snippet_start > 0 else ""
        suffix = "..." if snippet_end < len(text) else ""
        snippet = text[snippet_start:snippet_end]
        try:
            # استفاده از \g برای اطمینان از درج دقیق متن یافت شده (با حروف بزرگ/کوچک اصلی در متن)
            highlighted_snippet = re.sub(re.escape(phrase), r"**\g**", snippet, 1, re.IGNORECASE)
        except re.error as re_err:
            logging.warning(f"Regex error during highlighting phrase '{phrase}': {re_err}. Returning snippet without highlight.")
            highlighted_snippet = snippet

        return f"{prefix}{highlighted_snippet}{suffix}"
    except Exception as e:
        logging.error(f"Error generating snippet for phrase '{phrase}': {e}", exc_info=True)
        return "[خطا در ایجاد پیش‌نمایش]"


def analyze_phrases_in_text(
    phrases: List[str],
    page_text: Optional[str], # می‌تواند None باشد
    url: str
) -> List[Dict[str, Union[str, int, bool, List[str], None]]]:
    """Analyzes phrases against page text (exact phrase matching, case-insensitive)."""
    results = []
    if not page_text:
        logging.warning(f"Page text is empty or None for URL {url}. Returning default results for all phrases.")
        # ایجاد نتایج پیش‌فرض حتی اگر متن صفحه خالی باشد
        for phrase in phrases:
            if not phrase or not phrase.strip(): continue # رد کردن عبارات خالی
            phrase_lower = phrase.strip().lower()
            words_in_phrase = re.findall(r'\b\w+\b', phrase_lower)
            important_terms = [word for word in words_in_phrase if word not in ALL_STOP_WORDS]
            results.append({
                'original_phrase': phrase, 'phrase_to_find': phrase_lower,
                'found_phrase_count': 0, 'important_terms': important_terms,
                'total_score': 0, 'found_phrase': False,
                'analysis_notes': "محتوای صفحه وب قابل دریافت یا خالی بود.", 'url': url, 'context_snippet': None
            })
        return results

    logging.info(f"Analyzing {len(phrases)} phrases against text of length {len(page_text)} from {url}")
    page_text_lower = page_text # متن از قبل lowercase شده

    for phrase in phrases:
        if not phrase or not phrase.strip(): continue

        original_phrase = phrase
        phrase_lower = phrase.strip().lower()
        words_in_phrase = re.findall(r'\b\w+\b', phrase_lower)
        important_terms = [word for word in words_in_phrase if word not in ALL_STOP_WORDS]
        phrase_count = 0
        context_snippet = None
        analysis_notes = None

        try:
            escaped_phrase = re.escape(phrase_lower)
            matches = list(re.finditer(escaped_phrase, page_text_lower, re.IGNORECASE))
            phrase_count = len(matches)

            if phrase_count > 0:
                context_snippet = get_context_snippet(page_text_lower, phrase_lower, SNIPPET_CONTEXT_LENGTH)
            else:
                 analysis_notes = "عبارت مورد نظر در متن صفحه یافت نشد."
        except re.error as e:
            logging.error(f"Regex error searching for phrase '{phrase_lower}': {e}")
            analysis_notes = f"خطای Regex هنگام جستجو: {e}"
        except Exception as e:
             logging.error(f"Unexpected error during analysis of phrase '{phrase_lower}': {e}", exc_info=True)
             analysis_notes = f"خطای پیش‌بینی نشده هنگام تحلیل عبارت."

        total_score = phrase_count
        found_phrase = phrase_count > 0

        results.append({
            'original_phrase': original_phrase, 'phrase_to_find': phrase_lower,
            'found_phrase_count': phrase_count, 'important_terms': important_terms,
            'total_score': total_score, 'found_phrase': found_phrase,
            'analysis_notes': analysis_notes, 'url': url, 'context_snippet': context_snippet
        })

    found_count = sum(1 for r in results if r['found_phrase'])
    logging.info(f"Analysis complete for {url}. Found matches for {found_count} out of {len(results)} valid phrases.")
    return results


def generate_excel_report(results: List[Dict], url_checked: str) -> Optional[str]:
    """Generates a formatted Excel report from results and saves it."""
    if not results:
        logging.warning("No results provided to generate report.")
        return None

    report_data = []
    for res in results:
        important_terms_str = ', '.join(res.get('important_terms', [])) if res.get('important_terms') else "-"
        analysis_notes = res.get('analysis_notes') if res.get('analysis_notes') else ("یافت شد" if res.get('found_phrase') else "یافت نشد")
        context_snippet = res.get('context_snippet', "-")

        report_data.append({
            'Original Phrase': res.get('original_phrase', 'N/A'),
            'Phrase Searched (Lowercase)': res.get('phrase_to_find', 'N/A'),
            'Important Terms (Non-StopWords)': important_terms_str,
            'Times Found': res.get('found_phrase_count', 0),
            'Score (Phrase Count)': res.get('total_score', 0),
            'Phrase Found?': 'Yes' if res.get('found_phrase', False) else 'No',
            'Analysis Notes': analysis_notes,
            'Context Snippet': context_snippet,
            'URL Checked': url_checked
        })

    try:
        df = pd.DataFrame(report_data)
        output_name = f"analysis_report_{uuid.uuid4().hex[:8]}.xlsx"
        output_path = os.path.join(app.config['RESULTS_FOLDER'], output_name)
        sheet_name = 'Analysis Results' # نام شیت اکسل

        # استفاده از ExcelWriter برای دسترسی به قابلیت‌های openpyxl
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # گرفتن آبجکت‌های workbook و worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # تعریف استایل‌ها
            header_font = Font(bold=True)
            # تنظیم تراز برای متن (چپ و بالا) و اعداد (راست و وسط)
            text_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            number_alignment = Alignment(horizontal='right', vertical='center')

            # فرمت‌بندی سربرگ و تنظیم عرض ستون‌ها
            for col_num, column_title in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_num)
                cell = worksheet[f"{column_letter}1"] # سلول سربرگ

                # اعمال فونت بولد به سربرگ
                cell.font = header_font

                # محاسبه عرض ستون بر اساس طولانی‌ترین مقدار یا سربرگ
                max_length = 0
                # در نظر گرفتن طول سربرگ
                column_width_title = len(str(column_title))
                max_length = max(max_length, column_width_title)

                # پیدا کردن طولانی‌ترین مقدار در ستون (با یک محدودیت برای سرعت)
                for i, cell_value in enumerate(df[column_title]):
                     # بررسی فقط ۱۰۰۰ ردیف اول برای سرعت
                     if i > 1000: break
                     cell_len = len(str(cell_value))
                     if cell_len > max_length:
                         max_length = cell_len

                # تنظیم عرض ستون (با کمی اضافه و محدودیت حداکثر/حداقل)
                # عرض تقریبی بر اساس تعداد کاراکتر
                adjusted_width = min(max(max_length + 2, column_width_title + 2), 60) # حداقل عرض بر اساس سربرگ، حداکثر 60

                # ستون‌های خاص که نیاز به عرض بیشتر یا کمتر دارند
                if column_title == 'Context Snippet':
                    adjusted_width = 50 # عرض بیشتر برای کانتکست
                    wrap_text_cols = True
                elif column_title == 'Times Found' or column_title == 'Score (Phrase Count)':
                    adjusted_width = 15 # عرض کمتر برای اعداد
                    wrap_text_cols = False
                elif column_title == 'Phrase Found?':
                     adjusted_width = 15
                     wrap_text_cols = False
                else:
                     wrap_text_cols = True # بقیه ستون‌ها متن طولانی‌تری دارند

                worksheet.column_dimensions[column_letter].width = adjusted_width

                # اعمال تراز بندی و شکستن متن به سلول‌های داده (غیر از سربرگ)
                for row_num in range(2, worksheet.max_row + 1):
                    data_cell = worksheet[f"{column_letter}{row_num}"]
                    if column_title in ['Times Found', 'Score (Phrase Count)']:
                        data_cell.alignment = number_alignment
                    else:
                        # فعال کردن Wrap Text فقط برای ستون‌های مشخص شده
                        if wrap_text_cols:
                             data_cell.alignment = text_alignment
                        else:
                             # اگر wrap text نخواهیم، فقط تراز افقی چپ و عمودی وسط
                             data_cell.alignment = Alignment(horizontal='left', vertical='center')


            # ثابت کردن ردیف اول (Freeze Header Row)
            worksheet.freeze_panes = 'A2'

        logging.info(f"Generated formatted Excel report: {output_path} for URL: {url_checked}")
        return output_name

    except ImportError:
         logging.error("Cannot write Excel file. 'openpyxl' engine not installed? Run: pip install openpyxl")
         raise RuntimeError("خطا: وابستگی 'openpyxl' برای نوشتن فایل اکسل نصب نشده است.")
    except Exception as e:
        logging.error(f"Failed to generate formatted Excel report {output_path}: {e}", exc_info=True)
        raise RuntimeError(f"خطا در تولید فایل گزارش اکسل فرمت‌بندی شده: {e}") from e

    try:
        df = pd.DataFrame(report_data)
        output_name = f"analysis_report_{uuid.uuid4().hex[:8]}.xlsx"
        output_path = os.path.join(app.config['RESULTS_FOLDER'], output_name)

        df.to_excel(output_path, index=False, engine='openpyxl')
        logging.info(f"Generated Excel report: {output_path} for URL: {url_checked}")
        return output_name
    except ImportError:
         logging.error("Cannot write Excel file. 'openpyxl' engine not installed? Run: pip install openpyxl")
         raise RuntimeError("خطا: وابستگی 'openpyxl' برای نوشتن فایل اکسل نصب نشده است.")
    except Exception as e:
        logging.error(f"Failed to generate Excel report {output_path}: {e}", exc_info=True)
        # در محیط production بهتر است خطا را raise کنیم تا لاگ شود و 500 برگردد
        raise RuntimeError(f"خطا در تولید فایل گزارش اکسل: {e}") from e

# --- Flask Routes ---
@app.route('/', methods=['GET', 'POST'])
def index() -> str:
    """Handles file upload, website input, processing, and displaying results."""
    results_summary: Optional[List[Dict]] = None
    error: Optional[str] = None
    download_filename: Optional[str] = None
    processing_done: bool = False
    email: str = request.form.get('email', '').strip()
    website: str = request.form.get('website', '').strip()

    if request.method == 'POST':
        log_email = email if email else "Missing"
        log_website = website if website else "Missing"
        logging.info(f"POST request received. Email: {log_email}, Website: {log_website}")

        # --- Input Validation ---
        file = request.files.get('file')
        if not file or not file.filename:
            error = "لطفا فایل اکسل حاوی عبارات کلیدی را انتخاب کنید."
        elif not email:
            error = "لطفا ایمیل خود را وارد کنید."
        elif not validate_email_address(email):
              error = "فرمت ایمیل وارد شده صحیح نیست."
        elif not website:
            error = "لطفا آدرس وب‌سایت را وارد کنید."
        elif not allowed_file(file.filename):
            error = "فرمت فایل مجاز نیست. لطفا فایل اکسل با پسوند .xlsx آپلود کنید."

        # --- Processing ---
        if not error and file:
            website = validate_url(website)
            logging.info(f"Validated inputs. Processing website: {website}")
            filepath: Optional[str] = None
            try:
                filename = secure_filename(file.filename)
                unique_upload_name = f"{uuid.uuid4().hex[:8]}_{filename}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_upload_name)

                file.save(filepath)
                logging.info(f"File saved temporarily to {filepath}")

                if not is_valid_excel(filepath):
                    raise ValueError("فایل آپلود شده یک فایل اکسل معتبر (.xlsx) نیست.")

                phrases, phrases_error = get_phrases_from_file(filepath)
                if phrases_error: raise ValueError(f"خطا در پردازش فایل عبارات: {phrases_error}")
                if not phrases: raise ValueError("فایل اکسل عبارات معتبری ندارد.")

                logging.info(f"Fetching text content for {website}...")
                page_text, fetch_error = fetch_page_text(website)
                if fetch_error: raise ConnectionError(f"{fetch_error}")

                logging.info(f"Analyzing {len(phrases)} phrases...")
                analysis_results = analyze_phrases_in_text(phrases, page_text, website)

                if analysis_results:
                    logging.info("Generating Excel report...")
                    download_filename = generate_excel_report(analysis_results, website)
                    # generate_excel_report یا نام فایل را برمی‌گرداند یا خطا raise می‌کند
                    results_summary = analysis_results
                    logging.info(f"Excel report generated: {download_filename}")
                else:
                     error = "تحلیلی انجام نشد یا نتیجه‌ای حاصل نشد (ممکن است صفحه وب خالی بوده باشد)."
                     logging.warning("Analysis returned no results, possibly due to empty page text.")

                processing_done = True
                logging.info(f"Processing finished for {website}.")

            # مدیریت خطاهای مورد انتظار و نمایش پیام مناسب به کاربر
            except (ValueError, ConnectionError, RuntimeError, pd.errors.ParserError, ImportError) as e:
                error = str(e)
                logging.error(f"User-facing error during processing: {error}", exc_info=False) # لاگ کوتاه
            # مدیریت خطاهای ناشناخته (خطای 500 داخلی سرور)
            except Exception as e:
                error = f"یک خطای پیش‌بینی نشده در سرور رخ داد. لطفاً دوباره تلاش کنید یا با پشتیبانی تماس بگیرید."
                logging.error(f"Unexpected error during processing for {website}:", exc_info=True) # لاگ کامل
            finally:
                if filepath and os.path.exists(filepath):
                    try: os.remove(filepath)
                    except OSError as e: logging.error(f"Error removing uploaded file {filepath}: {e}")

    # رندر قالب HTML
    return render_template(
        "index.html",
        results=results_summary,
        error=error,
        download_filename=download_filename,
        processing_done=processing_done,
        email=email,
        website=website
    )

@app.route('/download/<filename>')
def download(filename: str) -> Response:
    """Provides the generated analysis Excel file for download."""
    safe_name = secure_filename(filename)
    if not safe_name or safe_name != filename:
         logging.warning(f"Download attempt with potentially unsafe filename rejected: Original='{filename}', Secured='{safe_name}'")
         return app.response_class("نام فایل نامعتبر است.", status=400, mimetype='text/plain')

    path = os.path.join(app.config['RESULTS_FOLDER'], safe_name)
    logging.info(f"Download request for file: {path}")

    if not os.path.isfile(path):
        logging.error(f"Download failed: File not found at {path}")
        return app.response_class("فایل مورد نظر یافت نشد.", status=404, mimetype='text/plain')

    try:
        return send_file(path, download_name=safe_name, as_attachment=True)
    except Exception as e:
         logging.error(f"Error sending file {path} for download: {e}", exc_info=True)
         return app.response_class("خطا در ارسال فایل.", status=500, mimetype='text/plain')

# --- Main Execution Guard ---
# !!! این بخش هرگز نباید در production اجرا شود !!!
# !!! سرور WSGI (مانند Gunicorn یا Waitress) مستقیماً شی 'app' را import و اجرا می‌کند !!!
if __name__ == '__main__':
    # فقط برای اجرای محلی جهت تست سریع
    print("اجرای برنامه در حالت توسعه (DEBUG MODE)...")
    print("هشدار: هرگز از این حالت برای production استفاده نکنید!")
    app.run(debug=True, host='0.0.0.0', port=5000) # برای تست محلی، این خط را از کامنت خارج کنید
    pass
